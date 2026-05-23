import os
import io
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)


def products_to_dataframe(products: List[Dict]) -> pd.DataFrame:
    rows = []
    for p in products:
        rows.append({
            "Site": p.get("site", ""),
            "Nom Produit": p.get("name", ""),
            "Marque": p.get("brand", "") or "",
            "Catégorie": p.get("category", "") or "",
            "Sous-catégorie": p.get("subcategory", "") or "",
            "Prix (TND)": p.get("price") or "",
            "Ancien Prix (TND)": p.get("old_price") or "",
            "Remise (%)": p.get("discount_percent") or "",
            "Promotion": "Oui" if p.get("has_promotion") else "Non",
            "En Stock": "Oui" if p.get("in_stock") else "Non",
            "URL": p.get("product_url", ""),
            "Image": p.get("image_url", "") or "",
        })
    return pd.DataFrame(rows)


def export_csv(products: List[Dict], filename: str = None) -> str:
    if not filename:
        filename = f"parabench_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = os.path.join(EXPORTS_DIR, filename)
    df = products_to_dataframe(products)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    return filepath


def export_excel(products: List[Dict], filename: str = None) -> str:
    if not filename:
        filename = f"parabench_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    
    df = products_to_dataframe(products)
    
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Main sheet
        df.to_excel(writer, sheet_name="Tous les Produits", index=False)
        
        # Per site sheets
        for site in ["parashop", "parafendri", "tunisiepara"]:
            site_df = df[df["Site"] == site]
            if not site_df.empty:
                site_df.to_excel(writer, sheet_name=site.capitalize(), index=False)
        
        # Promotions sheet
        promo_df = df[df["Promotion"] == "Oui"]
        if not promo_df.empty:
            promo_df.to_excel(writer, sheet_name="Promotions", index=False)
        
        # Styling
        workbook = writer.book
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Column widths
            col_widths = {
                "A": 15, "B": 50, "C": 25, "D": 20, "E": 20,
                "F": 12, "G": 12, "H": 10, "I": 10, "J": 10,
                "K": 60, "L": 60
            }
            for col, width in col_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Header style
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
    
    return filepath


def export_benchmark_excel(benchmark_data: List[Dict], filename: str = None) -> str:
    if not filename:
        filename = f"benchmark_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    
    rows = []
    for item in benchmark_data:
        row = {
            "Produit": item.get("product_name_normalized", ""),
            "Marque": item.get("brand", "") or "",
            "Prix Parashop (TND)": item.get("price_parashop") or "N/D",
            "Prix Parafendri (TND)": item.get("price_parafendri") or "N/D",
            "Prix TunisiePara (TND)": item.get("price_tunisiepara") or "N/D",
            "Prix Min (TND)": item.get("min_price") or "",
            "Prix Max (TND)": item.get("max_price") or "",
            "Différence (%)": item.get("price_diff_percent") or "",
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Benchmark Prix", index=False)
        
        workbook = writer.book
        worksheet = writer.sheets["Benchmark Prix"]
        
        from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
        header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        widths = {"A": 50, "B": 20, "C": 18, "D": 18, "E": 18, "F": 15, "G": 15, "H": 15}
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width
    
    return filepath


def export_pdf(products: List[Dict], filename: str = None) -> str:
    if not filename:
        filename = f"parabench_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(EXPORTS_DIR, filename)
    
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), 
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], 
                                  textColor=colors.HexColor("#1a1a2e"),
                                  fontSize=18, spaceAfter=20)
    
    elements = []
    elements.append(Paragraph("ParaBench — Export Produits", title_style))
    elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))
    
    # Table
    table_data = [["Site", "Produit", "Marque", "Catégorie", "Prix (TND)", "Promo", "Stock"]]
    
    for p in products[:500]:  # Limit for PDF
        table_data.append([
            p.get("site", ""),
            (p.get("name", "") or "")[:50],
            (p.get("brand", "") or "")[:20],
            (p.get("category", "") or "")[:20],
            str(p.get("price", "") or ""),
            "✓" if p.get("has_promotion") else "",
            "✓" if p.get("in_stock") else "✗",
        ])
    
    table = Table(table_data, repeatRows=1, 
                  colWidths=[3*cm, 8*cm, 4*cm, 4*cm, 3*cm, 2*cm, 2*cm])
    
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWHEIGHT", (0, 0), (-1, -1), 16),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return filepath
